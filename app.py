from flask import Flask, render_template, request, jsonify, redirect, url_for, send_file, session
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash
import os
import io
import json
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import psycopg2
import psycopg2.extras

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "cottage-secret-2024-change-me")

# ── Пользователи и роли ───────────────────────────────────
# Пароли можно переопределить переменными окружения на Railway.
USERS = {
    "admin": {
        "password": generate_password_hash(os.environ.get("APP_PASSWORD", "1234")),
        "role": "admin",
    },
    "services": {  # сотрудник по услугам
        "password": generate_password_hash(os.environ.get("SERVICES_PASSWORD", "services123")),
        "role": "services",
    },
    "reports": {   # сотрудник для отчётов
        "password": generate_password_hash(os.environ.get("REPORTS_PASSWORD", "reports123")),
        "role": "reports",
    },
}

# Набор прав каждой роли
ROLE_PERMS = {
    "admin":    {"bookings_view", "bookings_edit", "services_view", "services_edit",
                 "services_delete", "excel", "manage_users"},
    "services": {"services_view", "services_edit", "services_delete"},
    "reports":  {"services_view", "bookings_view", "excel"},
}


def current_role():
    return session.get("role")


def has_perm(perm):
    return perm in ROLE_PERMS.get(current_role(), set())


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("logged_in"):
            return redirect(url_for("login_page"))
        return f(*args, **kwargs)
    return decorated


def require_perm(perm):
    """Декоратор: требует конкретное право. Иначе 403 (или редирект на доступную страницу)."""
    def wrapper(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if not session.get("logged_in"):
                return redirect(url_for("login_page"))
            if not has_perm(perm):
                # Для API — 403 JSON, для страниц — редирект на стартовую доступную
                if request.path.startswith(("/export", "/cottages", "/bookings", "/service", "/settings", "/seed")) \
                   and request.method != "GET":
                    return jsonify({"error": "Недостаточно прав"}), 403
                if request.is_json or request.method != "GET":
                    return jsonify({"error": "Недостаточно прав"}), 403
                return redirect(url_for("home_redirect"))
            return f(*args, **kwargs)
        return decorated
    return wrapper


# ── База данных ───────────────────────────────────────────

def get_db():
    conn = psycopg2.connect(
        os.environ.get("DATABASE_URL", ""),
        cursor_factory=psycopg2.extras.RealDictCursor
    )
    return conn


def init_db():
    conn = get_db()
    cur  = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS cottages (
            id            SERIAL PRIMARY KEY,
            name          VARCHAR(255) NOT NULL,
            capacity      INT          DEFAULT 0,
            price_per_day FLOAT        DEFAULT 0,
            description   TEXT         DEFAULT '',
            contacts      TEXT         DEFAULT '',
            owner_type    VARCHAR(50)  DEFAULT 'Компания',
            property_type VARCHAR(50)  DEFAULT 'Коттедж',
            owner_name    VARCHAR(255) DEFAULT ''
        )
    """)
    # Миграции колонок
    cur.execute("ALTER TABLE cottages ADD COLUMN IF NOT EXISTS owner_type    VARCHAR(50)  DEFAULT 'Компания'")
    cur.execute("ALTER TABLE cottages ADD COLUMN IF NOT EXISTS property_type VARCHAR(50)  DEFAULT 'Коттедж'")
    cur.execute("ALTER TABLE cottages ADD COLUMN IF NOT EXISTS owner_name    VARCHAR(255) DEFAULT ''")
    cur.execute("ALTER TABLE cottages ADD COLUMN IF NOT EXISTS contacts      TEXT         DEFAULT ''")
    cur.execute("ALTER TABLE cottages ADD COLUMN IF NOT EXISTS capacity      INT          DEFAULT 0")
    cur.execute("ALTER TABLE cottages ADD COLUMN IF NOT EXISTS price_per_day FLOAT        DEFAULT 0")
    cur.execute("ALTER TABLE cottages ADD COLUMN IF NOT EXISTS cottage_size  VARCHAR(20)  DEFAULT ''")
    cur.execute("ALTER TABLE cottages ADD COLUMN IF NOT EXISTS rooms         INT          DEFAULT 0")
    cur.execute("ALTER TABLE cottages ADD COLUMN IF NOT EXISTS floor         VARCHAR(20)  DEFAULT ''")
    cur.execute("ALTER TABLE cottages ADD COLUMN IF NOT EXISTS position      INT          DEFAULT 0")
    cur.execute("UPDATE cottages SET position = id WHERE position = 0 OR position IS NULL")
    # Миграция данных: старые owner_type → новые owner_type + property_type
    cur.execute("UPDATE cottages SET owner_type='Компания',  property_type='Коттедж'     WHERE owner_type IN ('Алма-Ата','Коттеджи')")
    cur.execute("UPDATE cottages SET owner_type='Компания',  property_type='Номер отеля' WHERE owner_type = 'Номера отеля'")
    cur.execute("UPDATE cottages SET owner_type='Собственник'                            WHERE owner_type = 'Собственник' AND property_type IS NULL")
    cur.execute("UPDATE cottages SET property_type='Коттедж' WHERE property_type IS NULL OR property_type = ''")
    # Для собственников: перенести description → contacts (если contacts пусто).
    # Выполняется однократно — флаг в settings предотвращает повторный запуск,
    # который иначе затирал бы описание при каждом рестарте.
    cur.execute("SELECT value FROM settings WHERE key = 'mig_private_desc'")
    if not cur.fetchone():
        cur.execute("""
            UPDATE cottages
            SET contacts = description, description = ''
            WHERE owner_type = 'Собственник' AND (contacts IS NULL OR contacts = '') AND description != ''
        """)
        cur.execute("""
            INSERT INTO settings (key, value) VALUES ('mig_private_desc', 'done')
            ON CONFLICT (key) DO NOTHING
        """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS bookings (
            id                   SERIAL PRIMARY KEY,
            cottage_id           INT REFERENCES cottages(id) ON DELETE CASCADE,
            cottage_name         VARCHAR(255),
            guest_name           VARCHAR(255),
            guests               INT,
            check_in             DATE,
            check_out            DATE,
            nights               INT,
            discount             FLOAT   DEFAULT 0,
            total_before_discount FLOAT,
            total                FLOAT,
            rate                 FLOAT,
            total_som            FLOAT,
            notes                TEXT    DEFAULT ''
        )
    """)
    cur.execute("ALTER TABLE bookings ADD COLUMN IF NOT EXISTS deposit_paid FLOAT DEFAULT 0")
    cur.execute("ALTER TABLE bookings ADD COLUMN IF NOT EXISTS status VARCHAR(20) DEFAULT 'active'")
    cur.execute("ALTER TABLE bookings ADD COLUMN IF NOT EXISTS extra_per_night FLOAT DEFAULT 0")
    cur.execute("""
        CREATE TABLE IF NOT EXISTS settings (
            key   VARCHAR(50) PRIMARY KEY,
            value VARCHAR(255)
        )
    """)
    cur.execute("""
        INSERT INTO settings (key, value)
        VALUES ('rate', '85')
        ON CONFLICT (key) DO NOTHING
    """)

    # ── Однократная миграция валюты: $ → сомы ─────────────
    # Основная валюта теперь сом. Все денежные поля переводятся в сомы:
    #   • брони — по сохранённому в брони курсу (rate);
    #   • цены объектов — по текущему глобальному курсу.
    # Выполняется один раз (флаг som_migration_v1), иначе при каждом
    # рестарте суммы повторно умножались бы на курс.
    cur.execute("SELECT value FROM settings WHERE key = 'som_migration_v1'")
    if not cur.fetchone():
        cur.execute("SELECT value FROM settings WHERE key = 'rate'")
        _r = cur.fetchone()
        _global_rate = float(_r["value"]) if _r and _r["value"] else 85.0
        # Брони: total хранился в $, total_som = $ × курс. Переводим всё в сомы.
        cur.execute("""
            UPDATE bookings SET
                total                 = ROUND(COALESCE(total_som, total * COALESCE(rate,1))),
                total_before_discount = ROUND(COALESCE(total_before_discount,0) * COALESCE(rate,1)),
                discount              = ROUND(COALESCE(discount,0)              * COALESCE(rate,1)),
                deposit_paid          = ROUND(COALESCE(deposit_paid,0)          * COALESCE(rate,1)),
                extra_per_night       = ROUND(COALESCE(extra_per_night,0)       * COALESCE(rate,1))
            WHERE rate IS NOT NULL AND rate > 0
        """)
        # После перевода сумма в сомах совпадает с total
        cur.execute("UPDATE bookings SET total_som = total")
        # Цены объектов: $ → сомы по глобальному курсу
        cur.execute("UPDATE cottages SET price_per_day = ROUND(price_per_day * %s) WHERE price_per_day > 0",
                    (_global_rate,))
        cur.execute("INSERT INTO settings (key, value) VALUES ('som_migration_v1', 'done') ON CONFLICT (key) DO NOTHING")

    # ── Прайс-лист услуг ──────────────────────────────────
    cur.execute("""
        CREATE TABLE IF NOT EXISTS service_catalog (
            id         SERIAL PRIMARY KEY,
            category   VARCHAR(50)  NOT NULL,
            name       VARCHAR(255) NOT NULL,
            unit       VARCHAR(30)  DEFAULT 'шт',
            price      FLOAT        NOT NULL,
            has_plate  BOOLEAN      DEFAULT FALSE,
            owner_type VARCHAR(50)  DEFAULT 'Алма-Ата'
        )
    """)
    # Миграция: owner_type → канонические типы объектов
    cur.execute("ALTER TABLE service_catalog ADD COLUMN IF NOT EXISTS owner_type  VARCHAR(50) DEFAULT 'Коттедж'")
    cur.execute("ALTER TABLE service_catalog ADD COLUMN IF NOT EXISTS description TEXT        DEFAULT ''")
    cur.execute("UPDATE service_catalog SET owner_type = 'Коттедж'     WHERE owner_type IN ('Алма-Ата','Коттеджи')")
    cur.execute("UPDATE service_catalog SET owner_type = 'Номер отеля' WHERE owner_type = 'Номера отеля'")
    # Заполняем дефолтный прайс-лист (только если пустой)
    cur.execute("SELECT COUNT(*) as cnt FROM service_catalog")
    if cur.fetchone()["cnt"] == 0:
        defaults = [
            ("cleaning",    "Уборка — малый коттедж",          "шт",    6000, False),
            ("cleaning",    "Уборка — большой коттедж",        "шт",    8000, False),
            ("cleaning",    "Мытьё окон",                      "шт",    5000, False),
            ("cleaning",    "Уборка после ремонта",            "м²",     100, False),
            ("parking",     "Парковка (собственник)",          "сутки",   50, True),
            ("parking",     "Парковка (гость/арендатор)",      "сутки",  100, True),
            ("specialist",  "Вызов сантехника",                "шт",    2000, False),
            ("specialist",  "Вызов электрика",                 "шт",    2000, False),
            ("laundry",     "Простынь двуспальная Х/Б",        "шт",     400, False),
            ("laundry",     "Простынь малая Х/Б",              "шт",     250, False),
            ("laundry",     "Простынь махровая",               "шт",     500, False),
            ("laundry",     "Пододеяльник односпальный Х/Б",   "шт",     400, False),
            ("laundry",     "Наволочка Х/Б",                   "шт",     100, False),
            ("laundry",     "Полотенце большое",               "шт",     400, False),
            ("laundry",     "Полотенце среднее",               "шт",     300, False),
            ("laundry",     "Полотенце маленькое",             "шт",     200, False),
            ("laundry",     "Салфетка столовая",               "шт",     100, False),
            ("laundry",     "Халат махровый",                  "шт",     500, False),
            ("laundry",     "Халат Х/Б",                       "шт",     400, False),
            ("laundry",     "Домашний текстиль (шторы, тюль)", "кг",     800, False),
        ]
        cur.executemany(
            "INSERT INTO service_catalog (category, name, unit, price, has_plate) VALUES (%s,%s,%s,%s,%s)",
            defaults
        )

    # ── Журнал заказов услуг ──────────────────────────────
    cur.execute("""
        CREATE TABLE IF NOT EXISTS service_orders (
            id           SERIAL PRIMARY KEY,
            service_id   INT REFERENCES service_catalog(id) ON DELETE SET NULL,
            service_name VARCHAR(255),
            category     VARCHAR(50),
            cottage_id   INT REFERENCES cottages(id) ON DELETE SET NULL,
            cottage_name VARCHAR(255) DEFAULT '',
            service_date DATE        NOT NULL,
            end_date     DATE        DEFAULT NULL,
            quantity     FLOAT       DEFAULT 1,
            price        FLOAT       NOT NULL,
            total        FLOAT       NOT NULL,
            plate        VARCHAR(30) DEFAULT '',
            notes        TEXT        DEFAULT '',
            created_at   TIMESTAMP   DEFAULT NOW()
        )
    """)
    # Миграция: end_date + object_type (тип объекта на момент заказа)
    cur.execute("ALTER TABLE service_orders ADD COLUMN IF NOT EXISTS end_date DATE DEFAULT NULL")
    cur.execute("ALTER TABLE service_orders ADD COLUMN IF NOT EXISTS object_type VARCHAR(50) DEFAULT ''")
    cur.execute("ALTER TABLE service_orders ADD COLUMN IF NOT EXISTS position INT DEFAULT 0")
    cur.execute("UPDATE service_orders SET position = id WHERE position = 0 OR position IS NULL")
    # Бэкфилл object_type из связанного объекта
    cur.execute("""
        UPDATE service_orders so SET object_type = CASE
            WHEN c.owner_type = 'Собственник' THEN 'Собственник'
            ELSE c.property_type
        END
        FROM cottages c
        WHERE so.cottage_id = c.id AND (so.object_type IS NULL OR so.object_type = '')
    """)

    # ── Индексы для ускорения частых запросов ────────────
    # Брони почти всегда фильтруются/сортируются по объекту и датам,
    # услуги — по объекту и позиции. IF NOT EXISTS делает безопасным
    # повторный запуск при каждом старте.
    cur.execute("CREATE INDEX IF NOT EXISTS idx_bookings_cottage_checkin ON bookings (cottage_id, check_in)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_bookings_checkout        ON bookings (check_out)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_bookings_status          ON bookings (status)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_orders_cottage           ON service_orders (cottage_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_orders_position          ON service_orders (position)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_cottages_position        ON cottages (position)")

    conn.commit()
    cur.close()
    conn.close()


def fmt_date(iso) -> str:
    """YYYY-MM-DD или date → ДД/ММ/ГГГГ"""
    try:
        if isinstance(iso, date):
            return iso.strftime("%d/%m/%Y")
        return datetime.strptime(str(iso), "%Y-%m-%d").strftime("%d/%m/%Y")
    except Exception:
        return str(iso)


def serialize_booking(row) -> dict:
    """Конвертирует date-объекты PostgreSQL в строки ISO."""
    b = dict(row)
    for field in ("check_in", "check_out"):
        if isinstance(b.get(field), date):
            b[field] = b[field].isoformat()
    b["deposit_paid"]    = float(b.get("deposit_paid") or 0)
    b["extra_per_night"] = float(b.get("extra_per_night") or 0)
    b["status"]          = b.get("status") or "active"
    # При отмене брони остаток = 0 (аванс не возвращается, но больше ничего не взимается)
    if b["status"] == "cancelled":
        b["balance"] = 0.0
    else:
        b["balance"] = round(max(0.0, float(b.get("total") or 0) - b["deposit_paid"]), 2)
    return b


def get_rate(cur) -> float:
    cur.execute("SELECT value FROM settings WHERE key = 'rate'")
    row = cur.fetchone()
    return float(row["value"]) if row else 85.0


app.jinja_env.filters["fmtdate"] = fmt_date


@app.context_processor
def inject_perms():
    """Доступно во всех шаблонах: role, can('perm')."""
    return {"role": current_role(), "can": has_perm}


# ── Auth ──────────────────────────────────────────────────

@app.route("/login", methods=["GET", "POST"])
def login_page():
    if session.get("logged_in"):
        return redirect(url_for("home_redirect"))
    error = None
    if request.method == "POST":
        username = request.form.get("username", "")
        password = request.form.get("password", "")
        user = USERS.get(username)
        if user and check_password_hash(user["password"], password):
            session["logged_in"] = True
            session["role"]      = user["role"]
            session["username"]  = username
            return redirect(url_for("home_redirect"))
        error = "Неверный логин или пароль"
    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login_page"))


@app.route("/home")
@login_required
def home_redirect():
    """Отправляет пользователя на доступную ему стартовую страницу."""
    if has_perm("bookings_view"):
        return redirect(url_for("index"))
    if has_perm("services_view"):
        return redirect(url_for("services_page"))
    return redirect(url_for("logout"))


# ── Main ──────────────────────────────────────────────────

@app.route("/")
@require_perm("bookings_view")
def index():
    conn = get_db()
    cur  = conn.cursor()
    cur.execute("SELECT * FROM cottages ORDER BY position, id")
    cottages = [dict(r) for r in cur.fetchall()]
    # Счётчик активных (не прошедших и не отменённых) броней
    cur.execute("SELECT cottage_id, COUNT(*) AS cnt FROM bookings WHERE check_out >= CURRENT_DATE AND status != 'cancelled' GROUP BY cottage_id")
    counts = {r["cottage_id"]: r["cnt"] for r in cur.fetchall()}
    for c in cottages:
        c["bookings_count"] = counts.get(c["id"], 0)
    rate = get_rate(cur)
    cur.close(); conn.close()
    return render_template("index.html", cottages=cottages, rate=rate)


# Позиции меток объектов на карте (центр метки, % от размеров изображения).
# Координаты извлечены из текстового слоя исходного PDF-плана.
MAP_MARKERS = {
    'Т-2': (33.83, 55.04), 'Т-3': (29.57, 56.60), 'Т-5': (21.08, 59.12),
    'Б-1': (14.86, 61.51), 'Т-6': (20.00, 62.07), 'Б-13': (10.40, 62.46),
    'Б-12': (16.60, 64.24), 'Б-2': (11.64, 65.08), 'Б-11': (18.21, 66.44),
    'Б-3': (13.31, 67.53), 'Б-4': (14.86, 69.75), 'Б-10': (20.00, 69.02),
    'К-6': (23.95, 67.26), 'К-7': (27.32, 66.44), 'К-8': (30.84, 65.60),
    'К-12': (28.88, 69.02),
}


@app.route("/map")
@login_required
def map_page():
    """Опорный план комплекса с кликабельными объектами и подсветкой занятости."""
    conn = get_db(); cur = conn.cursor()
    # Сопоставляем метки карты с объектами по имени (без учёта лишних пробелов)
    cur.execute("SELECT id, name FROM cottages")
    name_to_id = {}
    for r in cur.fetchall():
        key = (r["name"] or "").strip()
        if key:
            name_to_id.setdefault(key, r["id"])
    # Занятые сегодня объекты (есть активная бронь, покрывающая текущую дату)
    cur.execute("""
        SELECT DISTINCT cottage_id FROM bookings
        WHERE status != 'cancelled' AND check_in <= CURRENT_DATE AND check_out > CURRENT_DATE
    """)
    occupied = {r["cottage_id"] for r in cur.fetchall()}
    cur.close(); conn.close()

    markers = []
    for label, (x, y) in MAP_MARKERS.items():
        cid = name_to_id.get(label)
        if cid is None:
            status = "none"          # объекта нет в системе
        elif cid in occupied:
            status = "occupied"      # занят сегодня
        else:
            status = "free"          # свободен
        markers.append({"label": label, "x": x, "y": y, "id": cid, "status": status})

    free_n     = sum(1 for m in markers if m["status"] == "free")
    occ_n      = sum(1 for m in markers if m["status"] == "occupied")
    return render_template("map.html", markers=markers, free_n=free_n, occ_n=occ_n)


# ── Settings ──────────────────────────────────────────────

@app.route("/settings", methods=["GET"])
@require_perm("bookings_view")
def get_settings():
    conn = get_db(); cur = conn.cursor()
    rate = get_rate(cur)
    cur.close(); conn.close()
    return jsonify({"rate": rate})


@app.route("/settings", methods=["POST"])
@require_perm("bookings_edit")
def update_settings():
    body = request.json
    conn = get_db(); cur = conn.cursor()
    if "rate" in body:
        cur.execute("UPDATE settings SET value = %s WHERE key = 'rate'", (str(body["rate"]),))
        conn.commit()
    rate = get_rate(cur)
    cur.close(); conn.close()
    return jsonify({"rate": rate})


# ── Cottages ──────────────────────────────────────────────

@app.route("/cottages", methods=["POST"])
@require_perm("bookings_edit")
def create_cottage():
    body = request.json
    conn = get_db(); cur = conn.cursor()
    owner_type    = body.get("owner_type", "Компания")
    property_type = body.get("property_type", "Коттедж")
    cur.execute("SELECT COALESCE(MAX(position),0)+1 AS p FROM cottages")
    pos = cur.fetchone()["p"]
    cur.execute("""
        INSERT INTO cottages (name, capacity, price_per_day, description, contacts,
                              owner_type, property_type, owner_name, cottage_size, rooms, floor, position)
        VALUES (%s,%s,%s,%s,%s, %s,%s,%s, %s,%s,%s, %s) RETURNING *
    """, (body["name"],
          int(body.get("capacity") or 0),
          float(body.get("price_per_day") or 0),
          body.get("description",""),
          body.get("contacts",""),
          owner_type, property_type,
          body.get("owner_name",""),
          body.get("cottage_size",""),
          int(body.get("rooms") or 0),
          body.get("floor",""), pos))
    cottage = dict(cur.fetchone())
    conn.commit(); cur.close(); conn.close()
    return jsonify(cottage), 201


@app.route("/cottages/<int:cottage_id>", methods=["PUT"])
@require_perm("bookings_edit")
def update_cottage(cottage_id):
    body = request.json
    conn = get_db(); cur = conn.cursor()
    owner_type    = body.get("owner_type", "Компания")
    property_type = body.get("property_type", "Коттедж")
    cur.execute("""
        UPDATE cottages SET name=%s, capacity=%s, price_per_day=%s,
                            description=%s, contacts=%s, owner_type=%s, property_type=%s, owner_name=%s,
                            cottage_size=%s, rooms=%s, floor=%s
        WHERE id=%s RETURNING *
    """, (body["name"],
          int(body.get("capacity") or 0),
          float(body.get("price_per_day") or 0),
          body.get("description",""),
          body.get("contacts",""),
          owner_type, property_type,
          body.get("owner_name",""),
          body.get("cottage_size",""),
          int(body.get("rooms") or 0),
          body.get("floor",""),
          cottage_id))
    row = cur.fetchone()
    conn.commit(); cur.close(); conn.close()
    if not row:
        return jsonify({"error": "Не найдено"}), 404
    return jsonify(dict(row))


@app.route("/cottages/<int:cottage_id>", methods=["DELETE"])
@require_perm("bookings_edit")
def delete_cottage(cottage_id):
    conn = get_db(); cur = conn.cursor()
    cur.execute("DELETE FROM cottages WHERE id = %s", (cottage_id,))
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True})


def _reorder(table, ids):
    """Перестановка строк, устойчивая к повторным перетаскиваниям и фильтрам.

    `ids` — желаемый порядок ПОДМНОЖЕСТВА строк (например, видимой вкладки).
    Алгоритм:
      1. читаем глобальный порядок всех строк (position, id);
      2. в «слоты», занятые подмножеством, подставляем его в новом порядке —
         относительный порядок остальных строк сохраняется;
      3. плотно перенумеровываем позиции 1..N.
    Плотная перенумерация исключает дубли/разрывы позиций, из-за которых
    при ORDER BY position,id порядок «слетал» после нескольких перестановок.
    """
    ids = [int(x) for x in ids]
    if not ids:
        return 0
    conn = get_db(); cur = conn.cursor()
    cur.execute(f"SELECT id FROM {table} ORDER BY position, id")
    all_ids = [r["id"] for r in cur.fetchall()]
    id_set  = set(ids)
    new_iter = iter(ids)
    result = []
    for oid in all_ids:
        result.append(next(new_iter) if oid in id_set else oid)
    # подстраховка: id из запроса, которых нет в таблице (рассинхрон) — в конец
    seen = set(result)
    for oid in ids:
        if oid not in seen:
            result.append(oid); seen.add(oid)
    for pos, oid in enumerate(result, 1):
        cur.execute(f"UPDATE {table} SET position = %s WHERE id = %s", (pos, oid))
    conn.commit(); cur.close(); conn.close()
    return len(ids)


@app.route("/cottages/reorder", methods=["POST"])
@require_perm("bookings_edit")
def reorder_cottages():
    ids = [int(x) for x in (request.json.get("ids") or [])]
    n = _reorder("cottages", ids)
    return jsonify({"ok": True, "updated": n})


@app.route("/service-orders/reorder", methods=["POST"])
@require_perm("services_edit")
def reorder_service_orders():
    ids = [int(x) for x in (request.json.get("ids") or [])]
    n = _reorder("service_orders", ids)
    return jsonify({"ok": True, "updated": n})


# ── Демо-данные ───────────────────────────────────────────

@app.route("/seed-demo", methods=["GET", "POST"])
@require_perm("bookings_edit")
def seed_demo():
    """Генерирует тестовые брони и заказы услуг для объектов компании."""
    from datetime import timedelta
    import random, traceback

    conn = get_db(); cur = conn.cursor()
    try:
        return _do_seed(cur, conn, timedelta, random)
    except Exception as e:
        conn.rollback(); cur.close(); conn.close()
        return jsonify({"ok": False, "error": str(e), "trace": traceback.format_exc()[-1500:]}), 500


def _do_seed(cur, conn, timedelta, random):
    # ── Демо-услуги для каждого типа объекта (если их нет) ──
    DEMO_SERVICES = {
        "Коттедж": [
            ("Уборка", "Генеральная уборка коттеджа", "шт", 7000),
            ("Стирка", "Комплект постельного белья", "компл", 1200),
            ("Сервис", "Доставка дров", "связка", 800),
        ],
        "Номер отеля": [
            ("Уборка", "Уборка номера", "шт", 1500),
            ("Сервис", "Завтрак в номер", "шт", 900),
            ("Стирка", "Смена полотенец", "компл", 500),
        ],
        "Квартира": [
            ("Уборка", "Уборка квартиры", "шт", 3000),
            ("Сервис", "Мытьё окон", "шт", 1500),
            ("Стирка", "Стирка штор", "кг", 700),
        ],
        "Номер для сотрудников": [
            ("Уборка", "Уборка комнаты", "шт", 1000),
            ("Стирка", "Смена белья", "компл", 600),
            ("Сервис", "Мелкий ремонт", "шт", 1500),
        ],
    }
    catalog_by_type = {}
    for otype, items in DEMO_SERVICES.items():
        cur.execute("SELECT id, name FROM service_catalog WHERE owner_type=%s", (otype,))
        existing = {r["name"] for r in cur.fetchall()}
        for cat, name, unit, price in items:
            if name not in existing:
                cur.execute("""
                    INSERT INTO service_catalog (category, name, unit, price, has_plate, owner_type)
                    VALUES (%s,%s,%s,%s,%s,%s)
                """, (cat, name, unit, price, False, otype))
        cur.execute("SELECT * FROM service_catalog WHERE owner_type=%s", (otype,))
        catalog_by_type[otype] = [dict(r) for r in cur.fetchall()]

    # ── Объекты компании ──
    cur.execute("SELECT * FROM cottages WHERE owner_type='Компания' ORDER BY id")
    objects = [dict(r) for r in cur.fetchall()]

    rate = get_rate(cur)
    guests_names = ["Азамат", "Нурлан", "Айгерим", "Данияр", "Жанна",
                    "Тимур", "Алина", "Бекзат", "Мадина", "Ерлан"]
    nights_variants = [1, 2, 3, 5, 7]

    bookings_added = 0
    orders_added   = 0

    for obj in objects:
        # Чистим старые демо-брони этого объекта чтобы не плодить дубли? Нет — просто добавляем без пересечений.
        # Определяем стартовую дату: после последней существующей брони
        cur.execute("SELECT MAX(check_out) AS m FROM bookings WHERE cottage_id=%s", (obj["id"],))
        row = cur.fetchone()
        cursor_date = (row["m"] if row and row["m"] else date.today())
        if isinstance(cursor_date, str):
            cursor_date = datetime.strptime(cursor_date, "%Y-%m-%d").date()
        cursor_date = max(cursor_date, date.today()) + timedelta(days=2)

        # Цена объекта теперь в сомах (после перехода на сомы)
        price = obj.get("price_per_day") or 30000
        shows_extra = obj["property_type"] in ('Коттедж', 'Номер отеля', 'Квартира')
        for i in range(5):
            nights = nights_variants[i % len(nights_variants)]
            ci = cursor_date
            co = ci + timedelta(days=nights)
            # Все суммы — в сомах
            extra_per_night = random.choice([0, 0, 0, 2000, 3000]) if shows_extra else 0
            discount        = random.choice([0, 0, 0, 2000, 5000])
            total_before    = round(nights * price + extra_per_night * nights)
            total           = max(0, total_before - discount)
            # Статус и аванс — демонстрируем новые поля
            if i == 0:
                status, deposit_paid = 'paid',      total
            elif i == 4:
                status, deposit_paid = 'cancelled', round(total * 0.2)
            else:
                status, deposit_paid = 'active',    random.choice([0, round(total * 0.3), round(total * 0.5)])
            cur.execute("""
                INSERT INTO bookings (cottage_id, cottage_name, guest_name, guests,
                    check_in, check_out, nights, discount, total_before_discount,
                    total, rate, total_som, notes, deposit_paid, extra_per_night, status)
                VALUES (%s,%s,%s,%s, %s,%s,%s, %s,%s,%s,%s,%s,%s, %s,%s,%s)
            """, (obj["id"], obj["name"], random.choice(guests_names),
                  max(1, (obj.get("capacity") or 2)), ci, co, nights,
                  discount, total_before, total, rate, total,
                  f"Демо-бронь #{i+1}", deposit_paid, extra_per_night, status))
            bookings_added += 1
            cursor_date = co + timedelta(days=random.randint(1, 3))  # зазор между бронями

        # ── 5 заказов услуг для типа этого объекта ──
        otype = obj["property_type"]
        services = catalog_by_type.get(otype, [])
        if services:
            for j in range(5):
                svc = random.choice(services)
                qty = random.choice([1, 1, 2, 3])
                sp  = svc["price"]
                sdate = date.today() + timedelta(days=random.randint(0, 20))
                cur.execute("""
                    INSERT INTO service_orders (service_id, service_name, category,
                        cottage_id, cottage_name, service_date, end_date, quantity,
                        price, total, plate, notes, object_type)
                    VALUES (%s,%s,%s,%s,%s, %s,%s,%s,%s,%s,%s,%s,%s)
                """, (svc["id"], svc["name"], svc["category"], obj["id"], obj["name"],
                      sdate, None, qty, sp, round(qty*sp), "",
                      f"Демо-заказ #{j+1}", otype))
                orders_added += 1

    conn.commit(); cur.close(); conn.close()
    return jsonify({
        "ok": True,
        "objects": len(objects),
        "bookings_added": bookings_added,
        "orders_added": orders_added,
        "message": f"Создано {bookings_added} броней и {orders_added} заказов услуг для {len(objects)} объектов."
    })


@app.route("/clear-demo", methods=["POST"])
@require_perm("bookings_edit")
def clear_demo():
    """Удаляет все записи, созданные через seed-demo."""
    conn = get_db(); cur = conn.cursor()
    cur.execute("DELETE FROM bookings WHERE notes LIKE 'Демо-бронь%'")
    b_count = cur.rowcount
    cur.execute("DELETE FROM service_orders WHERE notes LIKE 'Демо-заказ%'")
    o_count = cur.rowcount
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True, "message": f"Удалено {b_count} броней и {o_count} заказов."})


# ── Bookings ──────────────────────────────────────────────

@app.route("/bookings", methods=["GET"])
@require_perm("bookings_view")
def get_bookings():
    cottage_id = request.args.get("cottage_id", type=int)
    conn = get_db(); cur = conn.cursor()
    if cottage_id:
        cur.execute("SELECT * FROM bookings WHERE cottage_id=%s ORDER BY check_in", (cottage_id,))
    else:
        cur.execute("SELECT * FROM bookings ORDER BY check_in")
    rows = [serialize_booking(r) for r in cur.fetchall()]
    cur.close(); conn.close()
    return jsonify(rows)


@app.route("/bookings", methods=["POST"])
@require_perm("bookings_edit")
def create_booking():
    body       = request.json
    cottage_id = int(body["cottage_id"])
    check_in   = body["check_in"]
    check_out  = body["check_out"]
    guests     = int(body["guests"])

    conn = get_db(); cur = conn.cursor()

    cur.execute("SELECT * FROM cottages WHERE id = %s", (cottage_id,))
    cottage = cur.fetchone()
    if not cottage:
        cur.close(); conn.close()
        return jsonify({"error": "Коттедж не найден"}), 404

    # Лимит вместимости снимается, если указана доплата за доп. заселение
    if guests > cottage["capacity"] and float(body.get("extra_per_night") or 0) <= 0:
        cur.close(); conn.close()
        return jsonify({"error": f"Максимум гостей: {cottage['capacity']} (укажите доплату за доп. гостей, чтобы превысить)"}), 400

    ci = datetime.strptime(check_in,  "%Y-%m-%d").date()
    co = datetime.strptime(check_out, "%Y-%m-%d").date()
    if (co - ci).days < 2:
        cur.close(); conn.close()
        return jsonify({"error": "Минимальный срок брони — 2 ночи (посуточно не сдаём)"}), 400

    # Проверка пересечений (отменённые брони не блокируют даты)
    cur.execute("""
        SELECT id, check_in, check_out FROM bookings
        WHERE cottage_id = %s AND check_in < %s AND check_out > %s
          AND status != 'cancelled'
    """, (cottage_id, co, ci))
    conflict = cur.fetchone()
    if conflict:
        cur.close(); conn.close()
        return jsonify({"error": f"Даты пересекаются с бронью #{conflict['id']} ({fmt_date(conflict['check_in'])} – {fmt_date(conflict['check_out'])})"}), 409

    nights          = (co - ci).days
    extra_per_night = max(0, float(body.get("extra_per_night") or 0))      # сомы
    total_before    = nights * cottage["price_per_day"] + extra_per_night * nights  # сомы
    discount        = max(0, float(body.get("discount") or 0))             # сомы
    total           = round(max(0, total_before - discount))               # сомы
    rate            = get_rate(cur)        # глобальный курс — для пересчёта в $
    total_som       = total                # учёт ведётся в сомах
    deposit_paid    = max(0, float(body.get("deposit_paid") or 0))         # сомы

    cur.execute("""
        INSERT INTO bookings
            (cottage_id, cottage_name, guest_name, guests,
             check_in, check_out, nights,
             discount, total_before_discount, total, rate, total_som, notes, deposit_paid, extra_per_night)
        VALUES (%s,%s,%s,%s, %s,%s,%s, %s,%s,%s,%s,%s,%s,%s,%s)
        RETURNING *
    """, (cottage_id, cottage["name"], body.get("guest_name", ""), guests,
          ci, co, nights,
          discount, total_before, total, rate, total_som, body.get("notes", ""), deposit_paid, extra_per_night))
    booking = dict(cur.fetchone())
    conn.commit(); cur.close(); conn.close()

    # Сериализуем даты
    booking["check_in"]  = str(booking["check_in"])
    booking["check_out"] = str(booking["check_out"])
    return jsonify(booking), 201


@app.route("/bookings/<int:booking_id>", methods=["PUT"])
@require_perm("bookings_edit")
def update_booking(booking_id):
    body = request.json
    conn = get_db(); cur = conn.cursor()

    cur.execute("SELECT * FROM bookings WHERE id = %s", (booking_id,))
    existing = cur.fetchone()
    if not existing:
        cur.close(); conn.close()
        return jsonify({"error": "Бронь не найдена"}), 404

    # Коттедж: можно сменить объект
    cottage_id = int(body.get("cottage_id") or existing["cottage_id"])
    cur.execute("SELECT * FROM cottages WHERE id = %s", (cottage_id,))
    cottage = cur.fetchone()
    if not cottage:
        cur.close(); conn.close()
        return jsonify({"error": "Объект не найден"}), 404

    check_in  = body.get("check_in")  or str(existing["check_in"])
    check_out = body.get("check_out") or str(existing["check_out"])
    guests    = int(body.get("guests") or existing["guests"])

    # Эффективная доплата за доп. заселение (из запроса или из существующей брони)
    _extra_eff = body.get("extra_per_night")
    _extra_eff = float(_extra_eff) if _extra_eff is not None else float(existing.get("extra_per_night") or 0)
    # Лимит вместимости снимается, если указана доплата за доп. заселение
    if guests > cottage["capacity"] and _extra_eff <= 0:
        cur.close(); conn.close()
        return jsonify({"error": f"Максимум гостей: {cottage['capacity']} (укажите доплату за доп. гостей, чтобы превысить)"}), 400

    ci = datetime.strptime(check_in,  "%Y-%m-%d").date()
    co = datetime.strptime(check_out, "%Y-%m-%d").date()
    if (co - ci).days < 2:
        cur.close(); conn.close()
        return jsonify({"error": "Минимальный срок брони — 2 ночи (посуточно не сдаём)"}), 400

    # Проверка пересечений — исключая саму редактируемую бронь и отменённые
    cur.execute("""
        SELECT id, check_in, check_out FROM bookings
        WHERE cottage_id = %s AND id <> %s AND check_in < %s AND check_out > %s
          AND status != 'cancelled'
    """, (cottage_id, booking_id, co, ci))
    conflict = cur.fetchone()
    if conflict:
        cur.close(); conn.close()
        return jsonify({"error": f"Даты пересекаются с бронью #{conflict['id']} ({fmt_date(conflict['check_in'])} – {fmt_date(conflict['check_out'])})"}), 409

    nights          = (co - ci).days
    extra_per_night = max(0, float(body.get("extra_per_night") if body.get("extra_per_night") is not None else existing.get("extra_per_night") or 0))  # сомы
    total_before    = nights * cottage["price_per_day"] + extra_per_night * nights  # сомы
    discount        = max(0, float(body.get("discount") if body.get("discount") is not None else existing["discount"] or 0))  # сомы
    total           = round(max(0, total_before - discount))               # сомы
    rate            = float(existing["rate"] or get_rate(cur))   # курс для пересчёта в $
    total_som       = total                # учёт ведётся в сомах
    deposit_paid    = max(0, float(body.get("deposit_paid") if body.get("deposit_paid") is not None else existing.get("deposit_paid") or 0))  # сомы

    cur.execute("""
        UPDATE bookings SET cottage_id=%s, cottage_name=%s, guest_name=%s, guests=%s,
                            check_in=%s, check_out=%s, nights=%s, discount=%s,
                            total_before_discount=%s, total=%s, rate=%s, total_som=%s,
                            notes=%s, deposit_paid=%s, extra_per_night=%s
        WHERE id=%s RETURNING *
    """, (cottage_id, cottage["name"],
          body.get("guest_name", existing["guest_name"]), guests,
          ci, co, nights, discount, total_before, total, rate, total_som,
          body.get("notes", existing["notes"]), deposit_paid, extra_per_night, booking_id))
    booking = dict(cur.fetchone())
    conn.commit(); cur.close(); conn.close()
    booking["check_in"]  = str(booking["check_in"])
    booking["check_out"] = str(booking["check_out"])
    return jsonify(booking)


@app.route("/bookings/<int:booking_id>/status", methods=["PATCH"])
@require_perm("bookings_edit")
def update_booking_status(booking_id):
    new_status = request.json.get("status")
    if new_status not in ("active", "cancelled", "paid"):
        return jsonify({"error": "Недопустимый статус"}), 400
    conn = get_db(); cur = conn.cursor()
    cur.execute("UPDATE bookings SET status=%s WHERE id=%s RETURNING *", (new_status, booking_id))
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        return jsonify({"error": "Бронь не найдена"}), 404
    conn.commit(); cur.close(); conn.close()
    return jsonify(serialize_booking(row))


@app.route("/bookings/<int:booking_id>", methods=["DELETE"])
@require_perm("bookings_edit")
def delete_booking(booking_id):
    conn = get_db(); cur = conn.cursor()
    cur.execute("DELETE FROM bookings WHERE id = %s", (booking_id,))
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True})


@app.route("/bookings/bulk-delete", methods=["POST"])
@require_perm("bookings_edit")
def bulk_delete_bookings():
    ids = request.json.get("ids", [])
    if not ids:
        return jsonify({"ok": True, "deleted": 0})
    conn = get_db(); cur = conn.cursor()
    cur.execute("DELETE FROM bookings WHERE id = ANY(%s)", (ids,))
    count = cur.rowcount
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True, "deleted": count})


@app.route("/cottages/<int:cottage_id>/booked-ranges")
@require_perm("bookings_view")
def cottage_booked_ranges(cottage_id):
    """Занятые диапазоны дат объекта — для подсветки в календаре."""
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT check_in, check_out FROM bookings WHERE cottage_id = %s AND status != 'cancelled' ORDER BY check_in", (cottage_id,))
    ranges = [{"from": str(r["check_in"]), "to": str(r["check_out"])} for r in cur.fetchall()]
    cur.close(); conn.close()
    return jsonify(ranges)


@app.route("/cottages/<int:cottage_id>/bookings")
@require_perm("bookings_view")
def cottage_bookings_page(cottage_id):
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT * FROM cottages WHERE id = %s", (cottage_id,))
    cottage = cur.fetchone()
    if not cottage:
        cur.close(); conn.close()
        return redirect(url_for("index"))
    cur.execute("SELECT * FROM bookings WHERE cottage_id = %s ORDER BY check_in", (cottage_id,))
    bookings = [serialize_booking(r) for r in cur.fetchall()]
    rate  = get_rate(cur)
    today = date.today().isoformat()
    cur.close(); conn.close()
    # Занятые диапазоны для подсветки в календаре
    booked_ranges = [{"from": b["check_in"], "to": b["check_out"]} for b in bookings if b.get("status") != "cancelled"]
    # Итоговые суммы: для отменённых учитывается только аванс
    total_som = round(sum(_effective_total(b) for b in bookings))   # сомы (основная)
    total_usd = round(sum(_effective_usd(b)   for b in bookings))   # $ эквивалент
    return render_template("cottage.html", cottage=dict(cottage),
                           bookings=bookings, today=today, rate=rate,
                           booked_ranges=booked_ranges,
                           total_usd=total_usd, total_som=total_som)


# ── Services ──────────────────────────────────────────────

CATEGORY_LABELS = {
    "cleaning":   "🧹 Уборка",
    "parking":    "🚗 Парковка",
    "laundry":    "👕 Стирка",
    "specialist": "🔧 Специалист",
}

# Канонические типы объектов: значение БД → (ключ, подпись)
OBJECT_TYPES = [
    ("Коттедж",               "cottage",   "🏠 Коттеджи"),
    ("Номер отеля",           "hotel",     "🏨 Номера отеля"),
    ("Квартира",              "apartment", "🏢 Квартиры"),
    ("Номер для сотрудников", "employee",  "👷 Сотрудники"),
    ("Собственник",           "private",   "👤 Собственники"),
]
OBJ_KEY = {name: key for name, key, _ in OBJECT_TYPES}      # 'Коттедж' → 'cottage'
OBJ_LABEL = {name: lbl for name, _, lbl in OBJECT_TYPES}

@app.route("/services")
@require_perm("services_view")
def services_page():
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT * FROM service_catalog ORDER BY owner_type, category, id")
    catalog = [dict(r) for r in cur.fetchall()]

    # Группируем каталог по типу объекта → категории
    from collections import defaultdict
    grouped = defaultdict(lambda: defaultdict(list))
    for item in catalog:
        grouped[item["owner_type"]][item["category"]].append(item)
    # Каталоги по каждому типу
    catalogs = {key: dict(grouped.get(name, {})) for name, key, _ in OBJECT_TYPES}

    cur.execute("""
        SELECT so.*, c.name as cname, c.owner_type as c_owner, c.property_type as c_ptype
        FROM service_orders so
        LEFT JOIN cottages c ON c.id = so.cottage_id
        ORDER BY so.position, so.id
    """)
    orders = [dict(r) for r in cur.fetchall()]
    for o in orders:
        for f in ("service_date", "end_date"):
            if isinstance(o.get(f), date):
                o[f] = o[f].isoformat()
        # Тип объекта для фильтра: object_type или вычислить из объекта
        ot = o.get("object_type") or ""
        if not ot:
            if o.get("c_owner") == "Собственник":
                ot = "Собственник"
            elif o.get("c_ptype"):
                ot = o["c_ptype"]
        o["object_type"] = ot
        o["object_key"]  = OBJ_KEY.get(ot, "")
    cur.execute("SELECT * FROM cottages ORDER BY name")
    cottages = [dict(r) for r in cur.fetchall()]
    cur.close(); conn.close()
    return render_template("services.html",
        catalog=catalog, orders=orders,
        catalogs=catalogs, object_types=OBJECT_TYPES,
        cottages=cottages, categories=CATEGORY_LABELS,
        catalog_json=json.dumps(catalog))


@app.route("/service-catalog/<int:item_id>", methods=["PUT"])
@require_perm("services_edit")
def update_catalog_item(item_id):
    body = request.json
    conn = get_db(); cur = conn.cursor()
    cur.execute("""
        UPDATE service_catalog SET category=%s, name=%s, price=%s, unit=%s,
                                   description=%s, owner_type=%s
        WHERE id=%s RETURNING *
    """, (body.get("category","Услуга"), body["name"], float(body["price"]),
          body.get("unit","шт"), body.get("description",""),
          body.get("owner_type","Коттедж"), item_id))
    row = cur.fetchone()
    conn.commit(); cur.close(); conn.close()
    if not row: return jsonify({"error": "Не найдено"}), 404
    return jsonify(dict(row))


@app.route("/service-catalog", methods=["POST"])
@require_perm("services_edit")
def add_catalog_item():
    body = request.json
    conn = get_db(); cur = conn.cursor()
    cur.execute("""
        INSERT INTO service_catalog (category, name, unit, price, has_plate, owner_type, description)
        VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING *
    """, (body["category"], body["name"], body.get("unit","шт"),
          float(body["price"]), body.get("has_plate", False),
          body.get("owner_type","Коттедж"), body.get("description","")))
    row = dict(cur.fetchone())
    conn.commit(); cur.close(); conn.close()
    return jsonify(row), 201


@app.route("/service-catalog/<int:item_id>", methods=["DELETE"])
@require_perm("services_delete")
def delete_catalog_item(item_id):
    conn = get_db(); cur = conn.cursor()
    cur.execute("DELETE FROM service_catalog WHERE id=%s", (item_id,))
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True})


@app.route("/service-catalog/bulk-delete", methods=["POST"])
@require_perm("services_delete")
def bulk_delete_catalog():
    ids = request.json.get("ids", [])
    if not ids:
        return jsonify({"ok": True, "deleted": 0})
    conn = get_db(); cur = conn.cursor()
    cur.execute("DELETE FROM service_catalog WHERE id = ANY(%s)", (ids,))
    count = cur.rowcount
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True, "deleted": count})


@app.route("/service-orders", methods=["POST"])
@require_perm("services_edit")
def create_service_order():
    body = request.json
    conn = get_db(); cur = conn.cursor()

    svc = None
    if body.get("service_id"):
        cur.execute("SELECT * FROM service_catalog WHERE id=%s", (int(body["service_id"]),))
        svc = cur.fetchone()

    # Если услуга не из прайса — создаём «виртуальный» объект из переданных данных
    if not svc:
        custom_name = body.get("custom_name","").strip()
        if not custom_name:
            cur.close(); conn.close()
            return jsonify({"error": "Введите название услуги"}), 400
        svc = {"id": None, "name": custom_name, "category": "cleaning",
               "unit": "шт", "price": float(body.get("price",0)), "has_plate": False}

    cottage_id   = body.get("cottage_id") or None
    cottage_name = ""
    object_type  = body.get("object_type", "")   # тип объекта из формы
    if cottage_id:
        cur.execute("SELECT name, owner_type, property_type FROM cottages WHERE id=%s", (int(cottage_id),))
        c = cur.fetchone()
        if c:
            cottage_name = c["name"]
            if not object_type:
                object_type = "Собственник" if c["owner_type"] == "Собственник" else c["property_type"]

    price    = float(body.get("price") or svc["price"])
    end_date = body.get("end_date") or None

    # Для парковки считаем дни между датами автоматически
    if end_date and svc["has_plate"]:
        from datetime import date as _date
        d1 = datetime.strptime(body["service_date"], "%Y-%m-%d").date()
        d2 = datetime.strptime(end_date, "%Y-%m-%d").date()
        qty = max(1, (d2 - d1).days)
    else:
        qty = float(body.get("quantity") or 1)

    total = round(qty * price)

    # Используем кастомное название если введено вручную
    service_name = body.get("custom_name","").strip() or svc["name"]

    cur.execute("""
        INSERT INTO service_orders
            (service_id, service_name, category, cottage_id, cottage_name,
             service_date, end_date, quantity, price, total, plate, notes, object_type, position)
        VALUES (%s,%s,%s,%s,%s, %s,%s,%s,%s,%s,%s,%s,%s,
                (SELECT COALESCE(MAX(position),0)+1 FROM service_orders)) RETURNING *
    """, (svc["id"], service_name, svc["category"],
          cottage_id, cottage_name,
          body["service_date"], end_date, qty, price, total,
          body.get("plate",""), body.get("notes",""), object_type))
    order = dict(cur.fetchone())
    conn.commit(); cur.close(); conn.close()
    for f in ("service_date", "end_date"):
        if isinstance(order.get(f), date):
            order[f] = order[f].isoformat()
    return jsonify(order), 201


@app.route("/service-orders/<int:order_id>", methods=["PUT"])
@require_perm("services_edit")
def update_service_order(order_id):
    body = request.json
    conn = get_db(); cur = conn.cursor()

    cur.execute("SELECT * FROM service_orders WHERE id=%s", (order_id,))
    existing = cur.fetchone()
    if not existing:
        cur.close(); conn.close()
        return jsonify({"error": "Заказ не найден"}), 404

    svc = None
    if body.get("service_id"):
        cur.execute("SELECT * FROM service_catalog WHERE id=%s", (int(body["service_id"]),))
        svc = cur.fetchone()

    if not svc:
        custom_name = body.get("custom_name","").strip()
        if not custom_name:
            cur.close(); conn.close()
            return jsonify({"error": "Введите название услуги"}), 400
        svc = {"id": None, "name": custom_name, "category": existing["category"],
               "unit": "шт", "price": float(body.get("price") or existing["price"]),
               "has_plate": bool(existing["plate"])}

    cottage_id   = body.get("cottage_id") or None
    cottage_name = existing["cottage_name"] or ""
    object_type  = body.get("object_type") or existing.get("object_type") or ""
    if cottage_id:
        cur.execute("SELECT name, owner_type, property_type FROM cottages WHERE id=%s", (int(cottage_id),))
        c = cur.fetchone()
        if c:
            cottage_name = c["name"]
            if not object_type:
                object_type = "Собственник" if c["owner_type"] == "Собственник" else c["property_type"]

    price    = float(body.get("price") or svc["price"])
    end_date = body.get("end_date") or None

    if end_date and svc.get("has_plate"):
        d1 = datetime.strptime(body["service_date"], "%Y-%m-%d").date()
        d2 = datetime.strptime(end_date, "%Y-%m-%d").date()
        qty = max(1, (d2 - d1).days)
    else:
        qty = float(body.get("quantity") or 1)

    total        = round(qty * price)
    service_name = body.get("custom_name","").strip() or svc["name"]

    cur.execute("""
        UPDATE service_orders SET
            service_id=%s, service_name=%s, category=%s,
            cottage_id=%s, cottage_name=%s,
            service_date=%s, end_date=%s, quantity=%s,
            price=%s, total=%s, plate=%s, notes=%s, object_type=%s
        WHERE id=%s RETURNING *
    """, (svc["id"], service_name, svc["category"],
          cottage_id, cottage_name,
          body["service_date"], end_date, qty, price, total,
          body.get("plate",""), body.get("notes",""), object_type,
          order_id))
    order = dict(cur.fetchone())
    conn.commit(); cur.close(); conn.close()
    for f in ("service_date", "end_date"):
        if isinstance(order.get(f), date):
            order[f] = order[f].isoformat()
    return jsonify(order)


@app.route("/service-orders/<int:order_id>", methods=["DELETE"])
@require_perm("services_delete")
def delete_service_order(order_id):
    conn = get_db(); cur = conn.cursor()
    cur.execute("DELETE FROM service_orders WHERE id=%s", (order_id,))
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True})


@app.route("/service-orders/bulk-delete", methods=["POST"])
@require_perm("services_delete")
def bulk_delete_service_orders():
    ids = request.json.get("ids", [])
    if not ids:
        return jsonify({"ok": True, "deleted": 0})
    conn = get_db(); cur = conn.cursor()
    cur.execute("DELETE FROM service_orders WHERE id = ANY(%s)", (ids,))
    count = cur.rowcount
    conn.commit(); cur.close(); conn.close()
    return jsonify({"ok": True, "deleted": count})


def _write_services_to_sheet(ws, orders):
    """Записывает данные заказов услуг в существующий лист."""
    headers    = ["№","Дата начала","Дата окончания","Категория","Услуга",
                  "Коттеджи / Квартиры / Номера","Кол-во","Цена (сом)","Итого (сом)","Гос. номер","Заметки"]
    col_widths = [6, 14, 16, 18, 34, 30, 8, 13, 14, 16, 30]
    hf     = Font(bold=True, color="FFFFFF")
    hfill  = _header_fill("4F6EF7")
    border = _thin_border()
    center = Alignment(horizontal="center", vertical="center")
    CENTER = {1, 2, 3, 7, 8, 9}

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font=hf; cell.fill=hfill; cell.alignment=center; cell.border=border
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 22

    for seq, o in enumerate(orders, 1):
        ri = seq + 1
        values = [
            seq,
            fmt_date(o["service_date"]) if o.get("service_date") else "",
            fmt_date(o["end_date"])     if o.get("end_date")     else "—",
            o.get("category",""),
            o.get("service_name",""),
            o.get("cottage_name","") or "—",
            o.get("quantity", 1),
            o.get("price", 0),
            o.get("total", 0),
            o.get("plate","") or "—",
            o.get("notes",""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=ri, column=col, value=val)
            cell.fill = PatternFill("solid", fgColor="FFFFFF")
            cell.border = border
            if col in CENTER:
                cell.alignment = Alignment(horizontal="center")

    if orders:
        last = len(orders) + 2
        bold = Font(bold=True)
        fill = PatternFill("solid", fgColor="EEF2FF")
        for col in range(1, len(headers)+1):
            cell = ws.cell(row=last, column=col)
            cell.font=bold; cell.fill=fill; cell.border=border
        ws.cell(row=last, column=1, value="ИТОГО").font = bold
        total_cell = ws.cell(row=last, column=9, value=round(sum(o.get("total",0) for o in orders)))
        total_cell.font = bold
        total_cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:K{max(len(orders),1)+1}"


def _build_services_excel(orders, sheet_title):
    """Строит Excel-файл по списку заказов услуг."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]
    _write_services_to_sheet(ws, orders)
    return wb


def _fetch_service_orders(cur, object_type=None):
    """object_type — каноническое название типа ('Коттедж', 'Номер отеля', ...).
    Фильтрует по сохранённому object_type, а для старых записей — по типу связанного объекта."""
    if object_type:
        cur.execute("""
            SELECT so.*
            FROM service_orders so
            LEFT JOIN cottages c ON c.id = so.cottage_id
            WHERE COALESCE(NULLIF(so.object_type, ''),
                           CASE WHEN c.owner_type = 'Собственник' THEN 'Собственник'
                                ELSE c.property_type END) = %s
            ORDER BY so.position, so.id
        """, (object_type,))
    else:
        cur.execute("SELECT * FROM service_orders ORDER BY position, id")
    orders = [dict(r) for r in cur.fetchall()]
    for o in orders:
        for f in ("service_date", "end_date"):
            if isinstance(o.get(f), date):
                o[f] = o[f].isoformat()
    return orders


@app.route("/export/excel/services")
@require_perm("excel")
def export_excel_services():
    conn = get_db(); cur = conn.cursor()
    orders = _fetch_service_orders(cur)
    cur.close(); conn.close()
    wb  = _build_services_excel(orders, "Все услуги")
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"uslugi_vse_{date.today()}.xlsx")


@app.route("/export/excel/services/<obj_key>")
@require_perm("excel")
def export_excel_services_by_type(obj_key):
    """Экспорт услуг по типу объекта: cottage / hotel / apartment / employee / private."""
    KEY_TO_NAME = {key: name for name, key, _ in OBJECT_TYPES}
    obj_name = KEY_TO_NAME.get(obj_key)
    if not obj_name:
        return redirect(url_for("export_excel_services"))
    conn = get_db(); cur = conn.cursor()
    orders = _fetch_service_orders(cur, object_type=obj_name)
    cur.close(); conn.close()
    wb  = _build_services_excel(orders, f"Услуги — {obj_name}")
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"uslugi_{obj_key}_{date.today()}.xlsx")


# Названия листов без эмодзи (Excel не поддерживает эмодзи в именах листов)
_SVC_SHEET_TITLES = {
    "Коттедж":               "Коттеджи",
    "Номер отеля":           "Номера отеля",
    "Квартира":              "Квартиры",
    "Номер для сотрудников": "Сотрудники",
    "Собственник":           "Собственники",
}


@app.route("/export/excel/services-all")
@require_perm("excel")
def export_excel_services_all():
    """Единый Excel по всем услугам: отдельный лист на каждый тип объекта."""
    conn = get_db(); cur = conn.cursor()
    orders = _fetch_service_orders(cur)
    cur.close(); conn.close()

    wb = Workbook()

    # Первый лист: все услуги (включая собственников)
    ws = wb.active
    ws.title = "Все услуги"
    _write_services_to_sheet(ws, orders)

    # Дополнительные листы по типам объектов
    for obj_name, _key, _label in OBJECT_TYPES:
        type_orders = [o for o in orders if o.get("object_type") == obj_name]
        if not type_orders:
            continue
        ws2 = wb.create_sheet(_SVC_SHEET_TITLES.get(obj_name, obj_name)[:31])
        _write_services_to_sheet(ws2, type_orders)

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"uslugi_{date.today()}.xlsx")


# ── Excel helpers ─────────────────────────────────────────

BOOK_HEADERS = ["№","Коттеджи / Квартиры / Номера","Гость","Заезд","Выезд",
                "Ночей","Гостей","Скидка (сом)","Доп. гости (сом)","Сумма (сом)","Аванс (сом)","Сумма ($)","Статус","Заметки"]
BOOK_WIDTHS  = [6,30,22,13,13,8,8,13,16,14,13,12,12,30]
CENTER_COLS  = {1,6,7,8,9,10,11,12}

_STATUS_LABELS = {"active":"Активна","cancelled":"Отменена","paid":"Оплачена"}

def _effective_total(b):
    """Сумма в сомах. Для отменённых броней — только аванс (невозвратный)."""
    if b.get("status") == "cancelled":
        return float(b.get("deposit_paid") or 0)
    return float(b.get("total") or 0)

def _effective_usd(b):
    """Долларовый эквивалент суммы по курсу брони."""
    rate = float(b.get("rate") or 0)
    return round(_effective_total(b) / rate) if rate else 0

def _header_fill(color): return PatternFill("solid", fgColor=color)
def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _booking_row(b, seq_num):
    discount = round(b.get("discount") or 0)
    extra    = round((b.get("extra_per_night") or 0) * (b.get("nights") or 0))
    return [
        seq_num, b["cottage_name"], b["guest_name"],
        fmt_date(b["check_in"]), fmt_date(b["check_out"]),
        b["nights"], b["guests"],
        f"-{discount}" if discount else "—",
        extra if extra else "—",
        round(_effective_total(b)),          # Сумма (сомы)
        round(b.get("deposit_paid") or 0),   # Аванс (сомы)
        _effective_usd(b),                   # Сумма ($)
        _STATUS_LABELS.get(b.get("status") or "active", "Активна"),
        b.get("notes",""),
    ]

def _write_headers(ws, row=1):
    hf = Font(bold=True, color="FFFFFF")
    hfill = _header_fill("4F6EF7")
    border = _thin_border()
    center = Alignment(horizontal="center", vertical="center")
    for col,(h,w) in enumerate(zip(BOOK_HEADERS, BOOK_WIDTHS), 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font=hf; cell.fill=hfill; cell.alignment=center; cell.border=border
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[row].height = 22

def _write_rows(ws, bookings, start=2):
    today  = date.today().isoformat()
    border = _thin_border()
    last   = start - 1
    for seq, (row_i, b) in enumerate(zip(range(start, start + len(bookings)), bookings), 1):
        status   = b.get("status") or "active"
        is_past  = str(b["check_out"]) < today
        if status == "cancelled":
            row_fill = PatternFill("solid", fgColor="FFE4E4")   # розовый — отменена
        elif is_past:
            row_fill = PatternFill("solid", fgColor="F4F6F9")   # серый — прошедшая
        else:
            row_fill = PatternFill("solid", fgColor="FFFFFF")
        for col, val in enumerate(_booking_row(b, seq), 1):
            cell = ws.cell(row=row_i, column=col, value=val)
            cell.fill=row_fill; cell.border=border
            if col in CENTER_COLS:
                cell.alignment = Alignment(horizontal="center")
        last = row_i
    return last

def _write_totals(ws, bookings, total_row):
    border = _thin_border()
    bold   = Font(bold=True)
    fill   = PatternFill("solid", fgColor="EEF2FF")
    vals   = {1:"ИТОГО", 6:sum(b["nights"] for b in bookings),
              10:round(sum(_effective_total(b) for b in bookings)),
              11:round(sum((b.get("deposit_paid") or 0) for b in bookings)),
              12:round(sum(_effective_usd(b) for b in bookings))}
    for col in range(1, len(BOOK_HEADERS)+1):
        cell = ws.cell(row=total_row, column=col, value=vals.get(col))
        cell.font=bold; cell.fill=fill; cell.border=border
        if col in CENTER_COLS:
            cell.alignment = Alignment(horizontal="center")


@app.route("/export/excel")
@require_perm("excel")
def export_excel():
    """Единый Excel по броням: только объекты компании, отдельный лист на каждый тип."""
    conn = get_db(); cur = conn.cursor()
    # Исключаем собственников — только объекты компании
    cur.execute("SELECT * FROM cottages WHERE owner_type='Компания' ORDER BY position, id")
    company_cottages = [dict(r) for r in cur.fetchall()]
    cottage_ids = [c["id"] for c in company_cottages]

    if cottage_ids:
        cur.execute("SELECT * FROM bookings WHERE cottage_id = ANY(%s) ORDER BY check_in", (cottage_ids,))
        all_bookings = [serialize_booking(r) for r in cur.fetchall()]
    else:
        all_bookings = []
    cur.close(); conn.close()

    wb = Workbook()

    # Первый лист: все брони объектов компании
    ws = wb.active
    ws.title = "Все брони"
    _write_headers(ws, row=1)
    last = _write_rows(ws, all_bookings, start=2)
    if all_bookings: _write_totals(ws, all_bookings, last + 1)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:N{max(last, 2)}"

    # Дополнительные листы по типам объектов компании
    _BOOKING_TYPE_SHEETS = [
        ("Коттедж",               "Коттеджи"),
        ("Номер отеля",           "Номера отеля"),
        ("Квартира",              "Квартиры"),
        ("Номер для сотрудников", "Сотрудники"),
    ]
    for prop_type, sheet_title in _BOOKING_TYPE_SHEETS:
        type_ids      = {c["id"] for c in company_cottages if c["property_type"] == prop_type}
        type_bookings = [b for b in all_bookings if b["cottage_id"] in type_ids]
        if not type_bookings:
            continue
        ws2 = wb.create_sheet(sheet_title[:31])
        _write_headers(ws2, row=1)
        last2 = _write_rows(ws2, type_bookings, start=2)
        if type_bookings: _write_totals(ws2, type_bookings, last2 + 1)
        ws2.freeze_panes = "A2"
        ws2.auto_filter.ref = f"A1:N{max(last2, 2)}"

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name=f"broni_{date.today()}.xlsx")


@app.route("/export/excel/<int:cottage_id>")
@require_perm("excel")
def export_excel_cottage(cottage_id):
    conn=get_db(); cur=conn.cursor()
    cur.execute("SELECT * FROM cottages WHERE id=%s",(cottage_id,))
    cottage=cur.fetchone()
    if not cottage: cur.close();conn.close(); return jsonify({"error":"Не найдено"}),404
    cur.execute("SELECT * FROM bookings WHERE cottage_id=%s ORDER BY check_in",(cottage_id,))
    bookings=[serialize_booking(r) for r in cur.fetchall()]
    cur.close(); conn.close()

    wb=Workbook(); ws=wb.active; ws.title=cottage["name"][:31]
    ws.merge_cells("A1:N1"); tc=ws["A1"]
    tc.value=f"{cottage['name']}  |  до {cottage['capacity']} чел.  |  {int(cottage['price_per_day'])} сом/сутки"
    tc.font=Font(bold=True,size=12); tc.fill=PatternFill("solid",fgColor="EEF2FF")
    tc.alignment=Alignment(horizontal="left",vertical="center"); ws.row_dimensions[1].height=26
    _write_headers(ws, row=2)
    last=_write_rows(ws, bookings, start=3)
    if bookings: _write_totals(ws, bookings, last+1)
    ws.freeze_panes="A3"

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name=f"{cottage['name']}_{date.today()}.xlsx")


# ── Полный бэкап / восстановление (JSON) ──────────────────
# Все таблицы выгружаются в один файл .json; восстановление полностью
# заменяет текущие данные данными из файла (внутри транзакции).
# Порядок таблиц учитывает внешние ключи: родители раньше детей.
_BACKUP_TABLES = ["cottages", "service_catalog", "bookings", "service_orders", "settings"]


@app.route("/backup/download")
@require_perm("manage_users")
def backup_download():
    conn = get_db(); cur = conn.cursor()
    data = {}
    for t in _BACKUP_TABLES:
        cur.execute(f"SELECT * FROM {t} ORDER BY 1")
        data[t] = [dict(r) for r in cur.fetchall()]
    cur.close(); conn.close()
    payload = {"version": 1, "exported_at": date.today().isoformat(), "tables": data}
    # default=str безопасно сериализует даты/время/Decimal в строки
    raw = json.dumps(payload, ensure_ascii=False, default=str, indent=2)
    buf = io.BytesIO(raw.encode("utf-8"))
    return send_file(buf, mimetype="application/json", as_attachment=True,
                     download_name=f"backup_{date.today()}.json")


@app.route("/backup/restore", methods=["POST"])
@require_perm("manage_users")
def backup_restore():
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "Файл не выбран"}), 400
    try:
        payload = json.loads(f.read().decode("utf-8"))
    except Exception:
        return jsonify({"error": "Не удалось прочитать файл — это не похоже на JSON-бэкап"}), 400

    tables_data = payload.get("tables") if isinstance(payload, dict) else None
    if not isinstance(tables_data, dict) or "cottages" not in tables_data:
        return jsonify({"error": "Файл не похож на бэкап этой системы"}), 400

    conn = get_db(); cur = conn.cursor()
    try:
        # Чистим в порядке дети → родители
        for t in reversed(_BACKUP_TABLES):
            cur.execute(f"DELETE FROM {t}")
        # Вставляем в порядке родители → дети
        counts = {}
        for t in _BACKUP_TABLES:
            rows = tables_data.get(t) or []
            # реальные колонки таблицы (защита от инъекций и дрейфа схемы)
            cur.execute(f"SELECT * FROM {t} LIMIT 0")
            cols = [d[0] for d in cur.description]
            for row in rows:
                use = [c for c in cols if c in row]
                if not use:
                    continue
                collist = ",".join(use)
                placeholders = ",".join(["%s"] * len(use))
                cur.execute(f"INSERT INTO {t} ({collist}) VALUES ({placeholders})",
                            [row[c] for c in use])
            counts[t] = len(rows)
            # сбрасываем счётчик автоинкремента, чтобы новые id не конфликтовали
            if "id" in cols:
                cur.execute(
                    f"SELECT setval(pg_get_serial_sequence('{t}', 'id'), "
                    f"GREATEST((SELECT COALESCE(MAX(id), 0) FROM {t}), 1))"
                )
        conn.commit()
    except Exception as e:
        conn.rollback(); cur.close(); conn.close()
        return jsonify({"error": f"Ошибка восстановления, изменения отменены: {e}"}), 500
    cur.close(); conn.close()
    summary = ", ".join(f"{t}: {counts.get(t,0)}" for t in _BACKUP_TABLES)
    return jsonify({"ok": True, "message": "База восстановлена из бэкапа", "restored": summary})


# ── Старт ─────────────────────────────────────────────────

with app.app_context():
    try:
        init_db()
    except Exception as e:
        print(f"DB init skipped (no DATABASE_URL?): {e}")

@app.route("/export/excel/owner/<owner_type>")
@require_perm("excel")
def export_excel_by_owner(owner_type):
    MAP = {
        "company":   ("Коттеджи",       "owner_type='Компания' AND property_type='Коттедж'"),
        "cottage":   ("Коттеджи",       "owner_type='Компания' AND property_type='Коттедж'"),
        "hotel":     ("Номера отеля",   "owner_type='Компания' AND property_type='Номер отеля'"),
        "apartment": ("Квартиры",       "owner_type='Компания' AND property_type='Квартира'"),
        "employee":  ("Сотрудники",     "owner_type='Компания' AND property_type='Номер для сотрудников'"),
        "private":   ("Собственники",   "owner_type='Собственник'"),
    }
    if owner_type not in MAP:
        return redirect(url_for("export_excel"))
    label, where_clause = MAP[owner_type]

    # ── Собственники — справочник карточек, а не брони ────
    if owner_type == "private":
        conn = get_db(); cur = conn.cursor()
        cur.execute("SELECT * FROM cottages WHERE owner_type = 'Собственник' ORDER BY position, id")
        cottages = [dict(r) for r in cur.fetchall()]
        cur.close(); conn.close()

        wb = Workbook()
        ws = wb.active
        ws.title = "Собственники"

        headers    = ["№", "Название", "Тип объекта", "Собственник", "Контакты", "Описание"]
        col_widths = [6, 28, 18, 28, 40, 40]
        hf     = Font(bold=True, color="FFFFFF")
        hfill  = _header_fill("F59E0B")   # жёлтый акцент
        border = _thin_border()
        center = Alignment(horizontal="center", vertical="center")

        for col, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font=hf; cell.fill=hfill; cell.alignment=center; cell.border=border
            ws.column_dimensions[cell.column_letter].width = w
        ws.row_dimensions[1].height = 22

        for ri, c in enumerate(cottages, 2):
            row_fill = PatternFill("solid", fgColor="FFFFFF")
            values = [ri - 1, c["name"], c.get("property_type",""), c.get("owner_name",""), c.get("contacts",""), c.get("description","")]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=ri, column=col, value=val)
                cell.fill=row_fill; cell.border=border
                cell.alignment = Alignment(horizontal="center" if col==1 else "left", wrap_text=True)
            ws.row_dimensions[ri].height = 36

        ws.freeze_panes = "A2"
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return send_file(buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"sobstvenniki_{date.today()}.xlsx")

    # ── Коттеджи / Номера отеля — брони ───────────────────
    conn = get_db(); cur = conn.cursor()
    cur.execute(f"SELECT * FROM cottages WHERE {where_clause} ORDER BY position, id")
    cottages = [dict(r) for r in cur.fetchall()]
    cottage_ids = [c["id"] for c in cottages]

    if cottage_ids:
        cur.execute("SELECT * FROM bookings WHERE cottage_id = ANY(%s) ORDER BY check_in", (cottage_ids,))
        all_bookings = [serialize_booking(r) for r in cur.fetchall()]
    else:
        all_bookings = []
    cur.close(); conn.close()

    wb = Workbook()
    ws = wb.active; ws.title = f"Брони — {label}"
    _write_headers(ws, row=1)
    last = _write_rows(ws, all_bookings, start=2)
    if all_bookings: _write_totals(ws, all_bookings, last + 1)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:N{last}"

    for c in cottages:
        ws2 = wb.create_sheet(c["name"][:28])
        ws2.merge_cells("A1:N1"); tc = ws2["A1"]
        tc.value = f"{c['name']}  |  до {c['capacity']} чел.  |  {int(c['price_per_day'] or 0)} сом/сутки"
        tc.font  = Font(bold=True, size=12, color="2C3E50")
        tc.fill  = PatternFill("solid", fgColor="EEF2FF")
        tc.alignment = Alignment(horizontal="left", vertical="center")
        ws2.row_dimensions[1].height = 26
        _write_headers(ws2, row=2)
        cb   = [b for b in all_bookings if b["cottage_id"] == c["id"]]
        last = _write_rows(ws2, cb, start=3)
        if cb: _write_totals(ws2, cb, last + 1)
        ws2.freeze_panes = "A3"

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"broni_{label}_{date.today()}.xlsx")


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8080))
    # Отладчик включается только явно (FLASK_DEBUG=1) — в проде должен быть выключен
    app.run(host="0.0.0.0", port=port, debug=bool(os.environ.get("FLASK_DEBUG")))
